"""
DrawQuantPDF - Excel Exporter
Exports detection results and area calculations to professional Excel files.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from core.area_calculator import ProjectReport


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FACADE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FACADE_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
WINDOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
DOOR_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", size=11, bold=True, color="006100")
GRAND_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
GRAND_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

NUM_FMT = '#,##0.000'


class ExcelExporter:
    def __init__(self):
        self.wb = None
        self.ws = None

    def export(self, report: ProjectReport, output_path: str,
               project_name: str = "",
               cnp: str = "", beneficiar: str = "",
               localitate: str = "") -> str:
        """Export a ProjectReport to a formatted Excel file."""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Cantitati Fatade"

        self._write_title(project_name or report.project_name)
        row = self._write_headers(4)
        row = self._write_data(report, row)
        self._write_grand_totals(report, row)
        self._auto_column_widths()
        self._add_summary_sheet(report, project_name)
        self._add_quantities_sheet(report, cnp=cnp, beneficiar=beneficiar,
                                    localitate=localitate)

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        self.wb.save(output_path)
        return output_path

    def _write_title(self, project_name: str):
        ws = self.ws
        ws.merge_cells("A1:G1")
        title_cell = ws["A1"]
        title_cell.value = f"RAPORT CANTITATI - {project_name or 'DrawQuantPDF'}"
        title_cell.font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        ws.merge_cells("A2:G2")
        date_cell = ws["A2"]
        date_cell.value = f"Generat: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        date_cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
        date_cell.alignment = Alignment(horizontal="center")

    def _write_headers(self, start_row: int) -> int:
        headers = [
            ("Nr.", 6),
            ("Fatada", 25),
            ("Element", 20),
            ("Tip", 12),
            ("Latime (m)", 12),
            ("Inaltime (m)", 12),
            ("Suprafata (m²)", 16),
        ]

        for col, (header, width) in enumerate(headers, 1):
            cell = self.ws.cell(row=start_row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border = THIN_BORDER
            self.ws.column_dimensions[get_column_letter(col)].width = width

        self.ws.row_dimensions[start_row].height = 25
        return start_row + 1

    def _write_data(self, report: ProjectReport, row: int) -> int:
        nr = 1
        for facade in report.facades:
            row = self._write_facade_row(facade, row, nr)
            nr += 1

            for label, area in facade.windows:
                row = self._write_element_row(
                    row, facade.name, label, "Fereastra", area, WINDOW_FILL
                )

            for label, area in facade.doors:
                row = self._write_element_row(
                    row, facade.name, label, "Usa", area, DOOR_FILL
                )

            row = self._write_subtotal_row(facade, row)
        return row

    def _write_facade_row(self, facade, row: int, nr: int) -> int:
        ws = self.ws
        vals = [nr, facade.name, "FATADA", "Termosistem", "", "",
                facade.total_area_m2]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = FACADE_FONT
            cell.fill = FACADE_FILL
            cell.border = THIN_BORDER
            if col == 7 and isinstance(val, (int, float)):
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center" if col <= 4
                                           else "right")
        return row + 1

    def _write_element_row(self, row: int, facade_name: str, label: str,
                           tip: str, area: float, fill) -> int:
        ws = self.ws
        vals = ["", facade_name, label, tip, "", "", area]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.font = Font(name="Calibri", size=10)
            if col == 7 and isinstance(val, (int, float)):
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(
                    horizontal="center" if col <= 4 else "right"
                )
        return row + 1

    def _write_subtotal_row(self, facade, row: int) -> int:
        ws = self.ws
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell_label = ws.cell(row=row, column=1,
                             value=f"  Subtotal {facade.name}")
        cell_label.font = TOTAL_FONT
        cell_label.fill = TOTAL_FILL
        cell_label.alignment = Alignment(horizontal="right")
        cell_label.border = THIN_BORDER

        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = TOTAL_FILL
            ws.cell(row=row, column=col).border = THIN_BORDER

        cell_carp = ws.cell(row=row, column=6,
                            value=facade.total_carpentry_area)
        cell_carp.font = TOTAL_FONT
        cell_carp.fill = TOTAL_FILL
        cell_carp.number_format = NUM_FMT
        cell_carp.alignment = Alignment(horizontal="right")
        cell_carp.border = THIN_BORDER

        cell_net = ws.cell(row=row, column=7,
                           value=facade.net_thermosystem_area)
        cell_net.font = TOTAL_FONT
        cell_net.fill = TOTAL_FILL
        cell_net.number_format = NUM_FMT
        cell_net.alignment = Alignment(horizontal="right")
        cell_net.border = THIN_BORDER

        return row + 1

    def _write_grand_totals(self, report: ProjectReport, row: int):
        ws = self.ws
        row += 1

        totals = [
            ("TOTAL SUPRAFATA FATADE", report.total_facade_area),
            ("TOTAL TAMPLARIE (Ferestre + Usi)", report.total_carpentry_area),
            ("TOTAL TERMOSISTEM (Fatade - Tamplarie)",
             report.total_thermosystem_area),
            ("TOTAL SOCLU EXCLUS", report.total_socle_excluded_area_m2),
            ("TOTAL GLAF EXTERIOR", report.total_sill_length_m),
            ("TOTAL PICURATOR GOLURI", report.total_drip_profile_length_m),
            ("TOTAL LUNGIME COLTARE", report.total_corner_length_m),
            ("TOTAL PICURATOR SOCLU", report.total_socle_drip_profile_length_m),
            ("TOTAL PERIMETRU FERESTRE", report.total_window_perimeter_length_m),
            ("TOTAL PERIMETRU USI", report.total_door_perimeter_length_m),
        ]

        for label, value in totals:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=6)
            cell_label = ws.cell(row=row, column=1, value=label)
            cell_label.font = GRAND_FONT
            cell_label.fill = GRAND_FILL
            cell_label.alignment = Alignment(horizontal="right", vertical="center")
            cell_label.border = THIN_BORDER

            for col in range(2, 7):
                ws.cell(row=row, column=col).fill = GRAND_FILL
                ws.cell(row=row, column=col).border = THIN_BORDER

            cell_val = ws.cell(row=row, column=7, value=value)
            cell_val.font = GRAND_FONT
            cell_val.fill = GRAND_FILL
            cell_val.number_format = NUM_FMT
            cell_val.alignment = Alignment(horizontal="right", vertical="center")
            cell_val.border = THIN_BORDER

            ws.row_dimensions[row].height = 28
            row += 1

    def _auto_column_widths(self):
        for col in self.ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            adjusted = min(max_len + 4, 40)
            current = self.ws.column_dimensions[col_letter].width or 0
            if adjusted > current:
                self.ws.column_dimensions[col_letter].width = adjusted

    def _add_summary_sheet(self, report: ProjectReport, project_name: str):
        ws = self.wb.create_sheet("Sumar")

        ws.merge_cells("A1:K1")
        ws["A1"].value = "SUMAR PROIECT"
        ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = [
            "Fatada",
            "Sup. Totala (m²)",
            "Tamplarie (m²)",
            "Termosistem (m²)",
            "Soclu exclus (m²)",
            "Glaf exterior (m)",
            "Picurator goluri (m)",
            "Coltare (m)",
            "Picurator soclu (m)",
            "Perimetru ferestre (m)",
            "Perimetru usi (m)",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        row = 4
        for f in report.facades:
            vals = [
                f.name,
                f.total_area_m2,
                f.total_carpentry_area,
                f.net_thermosystem_area,
                f.socle_excluded_area_m2,
                f.sill_length_m,
                f.drip_profile_length_m,
                f.corner_length_m,
                f.socle_drip_profile_length_m,
                f.window_perimeter_length_m,
                f.door_perimeter_length_m,
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN_BORDER
                cell.font = Font(name="Calibri", size=10)
                if col > 1:
                    cell.number_format = NUM_FMT
                    cell.alignment = Alignment(horizontal="right")
            row += 1

        vals = [
            "TOTAL",
            report.total_facade_area,
            report.total_carpentry_area,
            report.total_thermosystem_area,
            report.total_socle_excluded_area_m2,
            report.total_sill_length_m,
            report.total_drip_profile_length_m,
            report.total_corner_length_m,
            report.total_socle_drip_profile_length_m,
            report.total_window_perimeter_length_m,
            report.total_door_perimeter_length_m,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = GRAND_FONT
            cell.fill = GRAND_FILL
            cell.border = THIN_BORDER
            if col > 1:
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")

        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 22

    def _add_quantities_sheet(self, report: ProjectReport, cnp: str = "",
                               beneficiar: str = "", localitate: str = ""):
        """Add 'Lista Cantitati' sheet — A4 portrait format with 4 sections."""
        ws = self.wb.create_sheet("Lista Cantitati")
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5

        # Column widths: A=description, B=mp, C=ml
        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14

        SECTION_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        LABEL_FONT = Font(name="Calibri", size=10)
        VALUE_FONT = Font(name="Calibri", size=10, bold=True)
        HEADER_BIG = Font(name="Calibri", size=14, bold=True, color="1F4E79")
        INFO_FONT = Font(name="Calibri", size=10, color="333333")
        TOTAL_BG = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        TOTAL_FNT = Font(name="Calibri", size=11, bold=True, color="006100")
        border = THIN_BORDER

        row = 1

        # ── HEADER ──
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value="LISTA CANTITĂȚI TERMOIZOLARE")
        c.font = HEADER_BIG
        c.alignment = Alignment(horizontal="center")
        row += 1

        for label, value in [("CNP:", cnp), ("Beneficiar:", beneficiar),
                              ("Localitate:", localitate)]:
            if value:
                ws.cell(row=row, column=1, value=label).font = INFO_FONT
                ws.cell(row=row, column=2, value=value).font = Font(
                    name="Calibri", size=10, bold=True)
                row += 1

        ws.cell(row=row, column=1,
                value=f"Generat: {datetime.now().strftime('%d.%m.%Y')}").font = Font(
                    name="Calibri", size=8, italic=True, color="999999")
        row += 2

        def section_header(title):
            nonlocal row
            for col in range(1, 4):
                c = ws.cell(row=row, column=col, value=title if col == 1 else "")
                c.font = SECTION_FONT
                c.fill = SECTION_FILL
                c.border = border
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            row += 1
            # Column sub-headers
            ws.cell(row=row, column=1, value="Descriere").font = LABEL_FONT
            ws.cell(row=row, column=2, value="mp").font = LABEL_FONT
            ws.cell(row=row, column=3, value="ml").font = LABEL_FONT
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
            row += 1

        def data_row(label, mp=None, ml=None):
            nonlocal row
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            ws.cell(row=row, column=1).border = border
            if mp is not None:
                c = ws.cell(row=row, column=2, value=round(mp, 2))
                c.font = VALUE_FONT
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=2).border = border
            if ml is not None:
                c = ws.cell(row=row, column=3, value=round(ml, 2))
                c.font = VALUE_FONT
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=3).border = border
            row += 1

        def total_row(label, mp=None, ml=None):
            nonlocal row
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = TOTAL_BG
                ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=1, value=label).font = TOTAL_FNT
            if mp is not None:
                c = ws.cell(row=row, column=2, value=round(mp, 2))
                c.font = TOTAL_FNT
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right")
            if ml is not None:
                c = ws.cell(row=row, column=3, value=round(ml, 2))
                c.font = TOTAL_FNT
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right")
            row += 1

        # ── SECTION 1: SUPRAFETE PRINCIPALE ──
        section_header("1. SUPRAFEȚE PRINCIPALE")
        data_row("Pod termoizolat", mp=0)  # placeholder — from floor plan (P1)
        data_row("Fațade total", mp=report.total_facade_area)
        data_row("Fațade fără goluri (termosistem)", mp=report.total_thermosystem_area)
        data_row("Vată minerală coșuri de fum", mp=0)  # manual input
        row += 1

        # ── SECTION 2: FERESTRE ──
        section_header("2. FERESTRE")
        data_row("Suprafață ferestre", mp=report.total_windows_area)
        data_row("Ferestre de înlocuit", mp=report.total_windows_area)
        data_row("Perimetru ferestre", ml=report.total_window_perimeter_length_m)
        data_row("Glaf exterior", ml=report.total_sill_length_m)
        data_row("Glaf interior", ml=report.total_sill_length_m)
        data_row("Colțare ferestre", ml=report.total_corner_length_m)
        data_row("Profil picurător ferestre", ml=report.total_sill_length_m)
        data_row("Șpaleți ferestre", mp=0)  # needs wall thickness
        row += 1

        # ── SECTION 3: USI ──
        section_header("3. UȘI")
        data_row("Suprafață uși", mp=report.total_doors_area)
        data_row("Uși de înlocuit", mp=report.total_doors_area)
        data_row("Perimetru uși", ml=report.total_door_perimeter_length_m)
        data_row("Colțare uși", ml=0)  # TODO: compute from door heights
        data_row("Profil picurător uși", ml=0)  # TODO
        data_row("Șpaleți uși", mp=0)  # needs wall thickness
        row += 1

        # ── SECTION 4: ELEMENTE PERIMETRALE ──
        section_header("4. ELEMENTE PERIMETRALE")
        data_row("Profil picurător soclu", ml=report.total_socle_drip_profile_length_m)
        data_row("Colțar exterior", ml=0)  # TODO: from facade geometry
        data_row("Finisaj interior", mp=0)  # TODO: perimeter × height
        row += 1

        # ── TOTAL ──
        total_row("TOTAL goluri (uși + ferestre)", mp=report.total_carpentry_area)
